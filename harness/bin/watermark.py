#!/usr/bin/env python3
"""JIVO watermark — every report that leaves this toolkit carries its origin.

Any HTML report or Excel workbook produced by an operator's Claude gets the
JIVO mark stamped into it: the brand badge, the operator's name and department,
and the date. Top-right, subtle, print-safe.

This tool NEVER touches SAP or any business system. It reads and writes only
the file it was handed. RULE 0 (read-only against every business system) is
untouched.

Subcommands
-----------
  stamp     Stamp one file (.html/.htm or .xlsx). Idempotent.
  hook      Read a Claude Code PostToolUse payload on stdin and stamp the
            file that was just written. This is the enforcement path.
  verify    Exit 1 if a file is missing its mark (for audit sweeps).
  whoami    Show the identity that would be stamped.

Design notes
------------
  * Enforcement is a HOOK, not a CLAUDE.md instruction. An instruction is a
    request the model can forget and an operator can talk it out of; a
    PostToolUse hook runs on every write, deterministically, with no model in
    the loop. `.claude/settings.json` is tracked in git, so this is live the
    moment someone clones — no install step for non-technical operators.
  * Identity NEVER fails open. If `harness/.operator` is missing we still
    stamp, using whatever the OS and git can tell us, tagged `unregistered`.
    A report with a weak attribution is recoverable; an unstamped report is
    not.
  * Stamping is idempotent and cheap to detect, so re-running a hook, or
    Claude editing a report five times, yields exactly one mark.
  * The badge is embedded as a base64 data URI, so a stamped HTML report is
    self-contained: it still shows the mark when mailed, copied to a USB
    stick, or opened on a machine that has never seen this repo.
"""

from __future__ import annotations

import argparse
import base64
import datetime as _dt
import getpass
import hashlib
import hmac
import json
import os
import re
import subprocess
import sys
from pathlib import Path

HARNESS = Path(__file__).resolve().parent.parent
REPO = HARNESS.parent
BADGE = HARNESS / "brand" / "daman-mark.png"
OPERATOR = HARNESS / ".operator"          # gitignored, per-machine
GUARD_HASH = HARNESS / "guard.hash"       # salted hash of the override password

MARK_VERSION = "v1"
HTML_MARKER = f"<!-- jivo-mark:{MARK_VERSION} -->"
XLSX_PROP = "JivoMark"

# The credit line. Fixed text, never an operator's name — it says who built the
# system that produced the report, not who ran it. Change it here and re-seal;
# it is deliberately not configurable per machine.
CREDIT = "powered by daman"

# Nobody's full legal name goes on a report. The operator line is a short
# lowercase first name, enough to trace a leaked report back to who produced it
# without turning every sheet into a nameplate.

# Stamp these. Everything else passes through untouched: .csv and .json are
# raw data with nowhere to put a mark, and .md reports render inside the chat
# rather than leaving the building.
HTML_SUFFIXES = {".html", ".htm"}
XLSX_SUFFIXES = {".xlsx", ".xlsm"}
STAMPABLE = HTML_SUFFIXES | XLSX_SUFFIXES


# ── identity ────────────────────────────────────────────────────────────────

def _git_config(key: str) -> str:
    try:
        out = subprocess.run(
            ["git", "config", "--get", key],
            cwd=REPO, capture_output=True, text=True, timeout=3,
        )
        return out.stdout.strip()
    except Exception:
        return ""


def _operator_identity() -> dict:
    """Who is producing this report.

    Resolution order: explicit env override, the registered operator file,
    then git, then the OS user. Never returns empty — an unattributable
    report is worse than a roughly-attributed one.
    """
    name = os.environ.get("JIVO_OPERATOR_NAME", "").strip()
    dept = os.environ.get("JIVO_OPERATOR_DEPT", "").strip()
    registered = True

    if not name and OPERATOR.exists():
        try:
            data = json.loads(OPERATOR.read_text(encoding="utf-8"))
            name = str(data.get("name", "")).strip()
            dept = str(data.get("department", "")).strip()
        except Exception:
            pass

    if not name:
        registered = False
        name = _git_config("user.name") or getpass.getuser() or "unknown"
    if not dept:
        # Department doubles as the harness persona when it wasn't set here.
        persona = HARNESS / ".persona"
        if persona.exists():
            dept = persona.read_text(encoding="utf-8").strip()
        dept = dept or "unassigned"

    return {
        "name": name,
        "department": dept,
        "registered": registered,
        "date": _dt.date.today().strftime("%d %b %Y"),
        "stamped_at": _dt.datetime.now().replace(microsecond=0).isoformat(),
    }


def _label(ident: dict) -> str:
    """The small operator line: first name, lowercase, plus the date."""
    first = ident["name"].strip().split()[0].lower() if ident["name"].strip() else "unknown"
    if not ident["registered"]:
        first += "?"
    return f"{first} · {ident['date']}"


# ── override gate ───────────────────────────────────────────────────────────

def _authorised(supplied: str | None) -> bool:
    """True only if `supplied` matches the stored override password.

    The password is never stored, shipped, or logged in plaintext — only a
    salted SHA-256. If no password has been set on this machine, there is no
    override at all and the answer is always no. Fail closed.
    """
    if not supplied or not GUARD_HASH.exists():
        return False
    try:
        raw = GUARD_HASH.read_text(encoding="utf-8").strip()
        salt_hex, want = raw.split("$", 1)
        got = hashlib.sha256(bytes.fromhex(salt_hex) + supplied.encode()).hexdigest()
        return hmac.compare_digest(got, want)
    except Exception:
        return False


# ── the badge ───────────────────────────────────────────────────────────────

def _badge_data_uri() -> str:
    if not BADGE.exists():
        return ""
    return "data:image/png;base64," + base64.b64encode(BADGE.read_bytes()).decode()


# ── HTML ────────────────────────────────────────────────────────────────────

def _html_block(ident: dict) -> str:
    uri = _badge_data_uri()
    img = (
        f'<img src="{uri}" alt="" width="30" height="30" '
        f'style="display:block;border-radius:7px;flex:0 0 auto">' if uri else ""
    )
    # No background plate: reports are documents, not dashboards, and a white
    # pill fights an editorial page. `currentColor` means the mark inherits the
    # page's own text colour, so it sits correctly on white, cream or dark
    # without needing a dark-mode branch.
    #
    # position:fixed repeats it on every page when printed or saved as PDF;
    # pointer-events:none keeps it from stealing clicks in a live report.
    return f"""{HTML_MARKER}
<div id="jivo-mark" aria-hidden="true">
<div style="text-align:right"><div>{_esc(CREDIT)}</div>
<div style="opacity:.6;font-size:9.5px;margin-top:1px">{_esc(_label(ident))}</div>
</div>{img}</div>
<style>
#jivo-mark{{position:fixed;top:16px;right:18px;z-index:2147483647;
 display:flex;gap:9px;align-items:center;pointer-events:none;opacity:.6;
 color:currentColor;letter-spacing:.01em;
 font:10.5px/1.3 ui-monospace,"IBM Plex Mono",SFMono-Regular,Menlo,monospace}}
@media print{{#jivo-mark{{position:fixed;opacity:.75;
 -webkit-print-color-adjust:exact;print-color-adjust:exact}}}}
</style>
"""


def _esc(s: str) -> str:
    return (s.replace("&", "&amp;").replace("<", "&lt;")
             .replace(">", "&gt;").replace('"', "&quot;"))


def _stamp_html(path: Path, ident: dict) -> str:
    text = path.read_text(encoding="utf-8", errors="surrogateescape")
    if HTML_MARKER in text:
        return "already-marked"

    block = _html_block(ident)
    # Insert before the LAST </body> so the mark cannot be pushed off-screen
    # by content appended afterwards.
    idx = text.lower().rfind("</body>")
    if idx != -1:
        out = text[:idx] + block + text[idx:]
    else:
        out = text + "\n" + block

    path.write_text(out, encoding="utf-8", errors="surrogateescape")
    return "stamped"


# ── XLSX ────────────────────────────────────────────────────────────────────

def _stamp_xlsx(path: Path, ident: dict) -> str:
    try:
        import openpyxl
        from openpyxl.drawing.image import Image as XLImage
        from openpyxl.packaging.custom import StringProperty
        from openpyxl.utils import get_column_letter
    except ImportError:
        return "skipped:openpyxl-missing"

    wb = openpyxl.load_workbook(path)

    try:
        if any(p.name == XLSX_PROP for p in wb.custom_doc_props.props):
            return "already-marked"
    except Exception:
        pass

    label = _label(ident)
    for ws in wb.worksheets:
        # Print header — this is what shows when the sheet is printed or
        # exported to PDF, which is how a workbook usually leaves a building.
        try:
            ws.oddHeader.right.text = f"{CREDIT} · {label}"
            ws.evenHeader.right.text = ws.oddHeader.right.text
        except Exception:
            pass

        # On-sheet badge, parked just right of the used range on row 1.
        if BADGE.exists():
            try:
                img = XLImage(str(BADGE))
                img.width = img.height = 56
                col = get_column_letter(min((ws.max_column or 1) + 2, 16384))
                img.anchor = f"{col}1"
                ws.add_image(img)
            except Exception:
                pass

    try:
        wb.custom_doc_props.append(
            StringProperty(name=XLSX_PROP, value=f"{MARK_VERSION}|{CREDIT}|{label}")
        )
    except Exception:
        pass

    wb.save(path)
    return "stamped"


# ── commands ────────────────────────────────────────────────────────────────

def _stamp_one(path: Path, ident: dict) -> str:
    suffix = path.suffix.lower()
    if suffix in HTML_SUFFIXES:
        return _stamp_html(path, ident)
    if suffix in XLSX_SUFFIXES:
        return _stamp_xlsx(path, ident)
    return "not-stampable"


def cmd_stamp(args: argparse.Namespace) -> int:
    if args.skip:
        if not _authorised(args.password or os.environ.get("JIVO_GUARD_PASSWORD")):
            print("watermark: --skip refused (no valid override password)",
                  file=sys.stderr)
            return 2
        print("watermark: skipped by authorised override")
        return 0

    ident = _operator_identity()
    rc = 0
    for raw in args.files:
        path = Path(raw)
        if not path.exists():
            print(f"watermark: no such file: {path}", file=sys.stderr)
            rc = 1
            continue
        if path.suffix.lower() not in STAMPABLE:
            if args.verbose:
                print(f"watermark: passthrough {path.name}")
            continue
        try:
            result = _stamp_one(path, ident)
        except Exception as exc:                       # never break the write
            print(f"watermark: {path.name}: {type(exc).__name__}: {exc}",
                  file=sys.stderr)
            rc = 1
            continue
        if args.verbose or result == "stamped":
            print(f"watermark: {result} {path.name}  [{_label(ident)}]")
    return rc


def cmd_hook(args: argparse.Namespace) -> int:
    """PostToolUse entry point. Stamps whatever Claude just wrote.

    Always exits 0. A hook that fails loudly on an unrelated write would
    break every session; the operator's work must not depend on this
    succeeding. Failures print to stderr, which Claude Code surfaces.
    """
    try:
        payload = json.load(sys.stdin)
    except Exception:
        return 0

    inp = payload.get("tool_input") or {}
    candidates = [inp.get("file_path"), inp.get("notebook_path"), inp.get("path")]
    for edit in (inp.get("edits") or []):
        if isinstance(edit, dict):
            candidates.append(edit.get("file_path"))

    ident = _operator_identity()
    for raw in candidates:
        if not raw:
            continue
        path = Path(str(raw))
        if path.suffix.lower() not in STAMPABLE or not path.exists():
            continue
        try:
            result = _stamp_one(path, ident)
            if result == "stamped":
                print(f"watermark: stamped {path.name} [{_label(ident)}]",
                      file=sys.stderr)
        except Exception as exc:
            print(f"watermark: {path.name}: {type(exc).__name__}: {exc}",
                  file=sys.stderr)
    return 0


def cmd_verify(args: argparse.Namespace) -> int:
    missing = []
    for raw in args.files:
        path = Path(raw)
        if not path.exists() or path.suffix.lower() not in STAMPABLE:
            continue
        marked = False
        if path.suffix.lower() in HTML_SUFFIXES:
            marked = HTML_MARKER in path.read_text(
                encoding="utf-8", errors="surrogateescape")
        else:
            try:
                import openpyxl
                wb = openpyxl.load_workbook(path)
                marked = any(p.name == XLSX_PROP for p in wb.custom_doc_props.props)
            except Exception:
                marked = False
        print(f"{'OK  ' if marked else 'MISS'} {path}")
        if not marked:
            missing.append(path)
    return 1 if missing else 0


def cmd_whoami(args: argparse.Namespace) -> int:
    ident = _operator_identity()
    print(json.dumps(ident, indent=2))
    # _label already carries the date; do not append it again.
    print(f"\nstamp: {CREDIT}\n       {_label(ident)}")
    if not ident["registered"]:
        print("\nNOT REGISTERED — run: python3 harness/bin/setup.py")
    return 0


def main() -> int:
    ap = argparse.ArgumentParser(
        prog="watermark",
        description="Stamp the JIVO mark onto reports leaving this toolkit.",
    )
    sub = ap.add_subparsers(dest="cmd", required=True)

    p = sub.add_parser("stamp", help="stamp one or more files")
    p.add_argument("files", nargs="*")
    p.add_argument("-v", "--verbose", action="store_true")
    p.add_argument("--skip", action="store_true",
                   help="suppress stamping (requires the override password)")
    p.add_argument("--password", help="override password (prefer the env var)")
    p.set_defaults(func=cmd_stamp)

    p = sub.add_parser("hook", help="PostToolUse hook entry point (stdin JSON)")
    p.set_defaults(func=cmd_hook)

    p = sub.add_parser("verify", help="report files missing their mark")
    p.add_argument("files", nargs="+")
    p.set_defaults(func=cmd_verify)

    p = sub.add_parser("whoami", help="show the identity that would be stamped")
    p.set_defaults(func=cmd_whoami)

    args = ap.parse_args()
    return args.func(args)


if __name__ == "__main__":
    sys.exit(main())
